from argparse import ONE_OR_MORE
from ast import Try
import asyncio
from importlib.resources import contents
from io import BytesIO
import json
import sqlite3
import os
import re
import random
from functools import partial
import aiogram
from aiogram import Bot, Dispatcher
from aiogram.types import Message
import datetime
from datetime import timedelta
import requests
import time
from time import sleep  
import telebot
from telebot import types
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery,ReplyKeyboardMarkup,ReplyKeyboardRemove
#import pandas as pd
from itertools import combinations, filterfalse
from Const import *
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import ttfonts
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
import openpyxl
import threading
import sched

def safe_paragraph(text):
    return Paragraph(str(text), styleN) if not isinstance(text, Paragraph) else text

class Notification:
    def __init__(self):
        self.text = None
        self.audio = None
        self.photo = None
        self.video = None
        self.document = None
        self.sticker = None
        self.video_note = None
        self.voice = None

    def send(self, bot, user_ids):
        for user_id in user_ids:
            try:
                print("Sending to ",user_id)
                if self.text:
                    bot.send_message(user_id, self.text)
                if self.audio:
                    bot.send_audio(user_id, self.audio)
                if self.photo:
                    bot.send_photo(user_id, self.photo)
                if self.video:
                    bot.send_video(user_id, self.video)
                if self.document:
                    bot.send_document(user_id, self.document)
                if self.sticker:
                    bot.send_sticker(user_id, self.sticker)
                if self.video_note:
                    bot.send_video_note(user_id, self.video_note)
                if self.voice:
                    bot.send_voice(user_id, self.voice)
            except Exception as e:
                print(f"Ошибка отправки пользователю {user_id}: {e}")

    def __del__(self):
        print(f"Obj {self} is deleted")


styles = getSampleStyleSheet()
styleN = styles["Normal"]



json_settings = open('/var/www/Bakery/Bakery.json','r') 
settings = json.load(json_settings)
TOKEN = settings['token']
db = settings['db']
arial = settings['arial']


pdfmetrics.registerFont(TTFont('Arial',arial))

user_order = {}

bot = telebot.TeleBot(TOKEN)

keys_cats = telebot.types.InlineKeyboardMarkup()

dp = Dispatcher()

msg = None

order_price = {}

scheduler = sched.scheduler(time.time, time.sleep)

@bot.message_handler(content_types=['document'], func=lambda message: message.from_user.id == 1159187641)
def doc(message: Message):
    global categories
    global dict_of_cats
    file_info = bot.get_file(message.document.file_id)
    file_name = message.document.file_name.lower()

    # Проверка расширения файла
    if file_name.endswith(('.xls', '.xlsx')):
        # Скачивание файла
        downloaded_file = bot.download_file(file_info.file_path)
        #file_path = os.path.join('downloads', message.document.file_name)

        # Сохранение файла на сервере
        file_path = '/var/www/Bakery/Catalog.xlsx'
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        categories = {}
        d = 0
        
            
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=4, values_only=True):
            
            category, product, detail1, detail2 = row
            
                    
            print("row ",row)
            detail1 = str(row[2])
            detail2 = str(row[3])
            if category not in categories and category is not None and category!='':
                print("Categorey: ",category)
                
                categories[category] = {}
                details = detail1 + ', '+ detail2
                categories[category][product] =str(d+1)+', '+ details
                
            elif category == '':
                continue
            elif category is None:
                continue
            elif category in categories:
                if product not in categories[category].keys():
                    
                    print("Product: ",product," appended")
                    categories[category][product] = str(d+1)+', '+ detail1 + ','+ detail2
                    
                            
            d = d+1
        workbook.close()
        c = 0
        for k in categories.keys():
            print("K: ",k)
            c+=1
            dict_of_cats[c]={}
            dict_of_cats[c][k]={}
            print("C: ",c)
            
            dict_of_cats[c][k] = categories[k]
        print("dict of cats: ",dict_of_cats)
                
                

async def main():
    updates = bot.get_updates()
    #id = updates[-1].message.chat.id
    
    con = sqlite3.connect(f'{db}')
    cur = con.cursor()
    
    cur.execute(f'''CREATE TABLE IF NOT EXISTS users(
                id INTEGER,
                name TEXT,
                ФИО TEXT,
                phone TEXT,
                order_text,
                start INTEGER,
                archive TEXT,
                questions TEXT,
                prepay TEXT
                )
                ''')
    con.commit()

    
    
    ids = cur.execute(f'''SELECT id FROM users''').fetchall()
    strt = telebot.types.ReplyKeyboardMarkup()
    s = telebot.types.KeyboardButton(text='/start')
    strt.add(s)
    #if ids:
    #    for i in ids:
            
    #        bot.send_message(i[0],'Бот был перезапущен, нажмите "/start"!',reply_markup=strt)
            
    #        bot.clear_step_handler_by_chat_id(i[0])
                
    try:
        starts = cur.execute(f'''UPDATE users SET start = {0} ''' )
        con.commit()
    except:
        pass
            

    cur.close()
    con.close()

    try:
        bot.infinity_polling()
    
    
    finally:
        bot.close()

categories = {}

admin_id = 183857728
#admin_id = 1159187641
nomen = None

all_orders = {}

normal_orders = {}

dict_of_cats = {}

@bot.message_handler(func=lambda message: message.text == '/update'  in message.text and message.from_user.id ==1159187641)
def upd(message:Message):
    con = sqlite3.connect(db)
    cur = con.cursor()
    users = cur.execute(f'''SELECT id FROM users WHERE id != {admin_id} ''').fetchall()
    for u in users:
        if u[0]!=466169926:
            bot.send_message(u[0],'Главное меню',reply_markup=client_markup())
        


    cur.close()
    con.close()

@bot.message_handler(func=lambda message: message.text == '/start' or '/start' in message.text)
def start(message:Message):
    
    user = message.from_user.id  
    
    con = sqlite3.connect(f'{db}') 
    cur = con.cursor()
    if con:
        print("Connection true")
    st = int()
    try:
        st = cur.execute(f'''SELECT start FROM users WHERE id ={user} ''').fetchone()[0]
    except:
        st = 0
    new_st = st+1
    print("This users start is: ",st)
    print("Next is: ",new_st)
    cur.execute('''SELECT id FROM users''') 
    result = cur.fetchall()
    print(result)
    ids = []
    id = message.chat.id
    for i in result:

        ids.append(i[0])
        
    if user in ids and user == admin_id:
        if st == 0:
            name = ''
            first = ''
            call_number = 1
            if message.chat.first_name:
                first = message.chat.first_name
            second = ''
            if message.chat.last_name:
            
                second = message.chat.last_name
            
            name = first + ' ' + second
            bot.send_message(user,f"Добро пожаловать, {name}!",reply_markup = ReplyKeyboardRemove())
            bot.send_message(user,f"Главное меню",reply_markup=admin_markup()) ###СДЕЛАТЬ МАРКАП
            
            
            
            bot.clear_step_handler_by_chat_id(user)
            #bot.register_next_step_handler_by_chat_id(user, admin_menu)
        else:
            cur.execute(f'''UPDATE users SET start = {new_st} WHERE id = {user} ''')
            con.commit()
            bot.send_message(user,f"Вы уже вводили старт), {name}!",reply_markup = ReplyKeyboardRemove())
            #time.sleep(0.)
            
            bot.send_message(admin_id,text='Главное меню') ###ДОБАВИТЬ МАРКАП
                
            
       
            bot.clear_step_handler_by_chat_id(admin_id)
        
    elif user not in ids:
        
        name = ''
        first = ''
        if message.chat.first_name:
            first = message.chat.first_name
        second = ''
        if message.chat.last_name:
            second = message.chat.last_name
            
        name = first + second
        if first != '':
            bot.send_message(user,f'Доброго времени суток, {first}.')
        elif first == '':
            bot.send_message(user,f'Доброго времени суток!')
            

        cur.execute(f'''INSERT INTO users (id,start) VALUES (?,?)''',(user,1,))
        con.commit()
        
        
        
        #time.sleep(0.5)
        bot.send_message(user,'Кофеек-Кренделёк приветствует вас!\nНажмите на кнопку "Поделиться контактом", чтобы отправить мне ваш телефон. Я вас зарегистрирую)',reply_markup=contact)
        bot.clear_step_handler_by_chat_id(user)
        bot.register_next_step_handler_by_chat_id(user, partial(wait_phone_number))
        
    elif user in ids and user != admin_id:
        cur.execute(f'''UPDATE users SET start = {new_st} WHERE id = {user} ''')
        con.commit()
        if st ==0:
        
            bot.send_message(user,f'Привет, {message.chat.first_name}! Рады тебя видеть)',reply_markup=ReplyKeyboardRemove())
            bot.send_message(user,f'Главное меню',reply_markup=client_markup())
            bot.clear_step_handler_by_chat_id(user)
            #time.sleep(0.5)
            

            cur.execute(f'''UPDATE users SET start = {new_st} WHERE id = {user} ''')
            con.commit()
        else:
            bot.send_message(user,f'Вы уже ввели старт)')
            
            
            bot.send_message(user,text='Главное меню',reply_markup=client_markup())
            bot.clear_step_handler_by_chat_id(user)
            #bot.register_next_step_handler_by_chat_id(user,partial(free_talk))

    cur.close()
    con.close()

        



def wait_phone_number(msg:Message):
    user = msg.from_user.id
    numbers =['1','2','3','4','5','6','7','8','9','0']
    try:
        
        phone = msg.contact.phone_number
        print("Phone: ",phone)
        phone = phone.replace("+7", "8")

        phone = "8" + phone[1:]
        print("Phone: ", phone)
        phone = phone.strip()
        phone = phone.replace('-','') 
        print("Is digit")
        print(phone)
        #phone = "8" + phone[1:]
        phone = phone.replace("+7", "8")
        digits = True
        for i in phone:
            if i in numbers:
                print(i, 'in numbers')
                continue
            else:
                digits = False
        if digits == True:
            print("yes")
            try:
                con = sqlite3.connect(f'{db}')
                cur = con.cursor()
            
                if user != admin_id:
                    name = msg.chat.first_name
                    if msg.chat.last_name:
                        full_name = name +" "+ msg.chat.last_name
                    else:
                        full_name = name
                    cur.execute(f'''UPDATE users SET phone = '{phone}' WHERE id = {user} ''')
                    print("Updating users setting phone = ",phone)
                    con.commit()
                    #bot.send_message(chat_id=user, text="Вы зарегистрированы в базе данных"+'\n'+"Выберите вашу лигу.",reply_markup=leagues)
                    #bot.clear_step_handler_by_chat_id(user)
                    #bot.register_next_step_handler_by_chat_id(user,partial(league))
                    bot.send_message(chat_id=user, text="Вы зарегистрированы в базе данных"+'\n'+"Введите ваши фамилию и имя.",reply_markup=ReplyKeyboardRemove())
                    bot.clear_step_handler_by_chat_id(user)
                    bot.register_next_step_handler_by_chat_id(user,partial(fio))
                if user == admin_id:
                    name = msg.chat.first_name
                    if msg.chat.last_name:
                        full_name = name +" "+ msg.chat.last_name
                    else:
                        full_name = name
                    cur.execute('''INSERT INTO users (id,name,phone) VALUES(?,?,?) ''', (user,full_name, phone,))
                    con.commit()
                    bot.send_message(chat_id=msg.chat.id, text=f"Добро пожаловать, {full_name})",reply_markup=admin_markup())
                    bot.clear_step_handler_by_chat_id(admin_id)
                   # bot.register_next_step_handler_by_chat_id(admin_id,menu)
                
                print("Executed")
                con.commit()
           
                con.close()
            except Exception as e:
                bot.send_message(user,text=f'Возникла ошибка: {e}.')
                
                bot.clear_step_handler_by_chat_id(user)
                bot.register_next_step_handler_by_chat_id(user,partial(wait_phone_number))
                try:
                    con.close()
                except:
                    pass
        else:
            markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
            contacts = types.KeyboardButton("Поделиться контактом",request_contact=True)
            markup.add(contacts)
            bot.send_message(user,text='Неправильный формат номера телефона. Попробуйте еще раз.',reply_markup=markup)
            bot.clear_step_handler_by_chat_id(user)
            bot.register_next_step_handler_by_chat_id(user, partial(wait_phone_number))
    except:
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
        contacts = types.KeyboardButton("Поделиться контактом",request_contact=True)
        markup.add(contacts)
        bot.send_message(chat_id=msg.chat.id,text="Пожалуйста, нажмите на кнопку.",reply_markup=markup)
        bot.clear_step_handler_by_chat_id(msg.chat.id)
        bot.register_next_step_handler_by_chat_id(msg.chat.id, partial(wait_phone_number))
    
    
 
        
def fio(message:Message):
    con = sqlite3.connect(f'{db}')
    cur = con.cursor()
    user = message.from_user.id
    n = message.text
    confirm_fio = telebot.types.InlineKeyboardMarkup()
    yes_fio = telebot.types.InlineKeyboardButton(text='Да, отправить ',callback_data='yep')
    no_fio = telebot.types.InlineKeyboardButton(text='Нет, изменить ',callback_data='nope')
    confirm_fio.add(yes_fio,no_fio)
    cur.execute(f'''UPDATE users SET ФИО = '{n}' WHERE id = {user} ''')
    con.commit()
    bot.send_message(user,f'Отправить данные: {n}?\n(Используйте клавиатуру бота для ответа){random.choice(bot_emojis)}',reply_markup=confirm_fio)
    bot.clear_step_handler_by_chat_id(user)
    bot.register_next_step_handler_by_chat_id(user,partial(confirm_this_fio))

    cur.close()
    con.close()

@bot.callback_query_handler(func=lambda call:True and call.data =='yep' or call.data =='nope')
def confirm_this_fio(call:CallbackQuery):
    con = sqlite3.connect(f'{db}')
    cur = con.cursor()
    user = call.from_user.id
    if call.data =='yep':
        bot.send_message(user,f'Отлично!\nНаслаждайтесь функциями бота) {random.choice(bot_emojis)}',reply_markup=ReplyKeyboardRemove())
        
        bot.send_message(user,f'Главное меню',reply_markup=client_markup())
        #bot.clear_step_handler_by_chat_id(user)
        #bot.register_next_step_handler_by_chat_id(user,partial(league))
        bot.clear_step_handler_by_chat_id(user)
        #bot.register_next_step_handler_by_chat_id(user,partial(menu))

    elif call.data == 'nope':
        bot.send_message(user,f'Хорошо, измените данные',reply_markup=ReplyKeyboardRemove())
        bot.clear_step_handler_by_chat_id(user)
        bot.register_next_step_handler_by_chat_id(user,partial(fio))

    else:
        bot.clear_step_handler_by_chat_id(user)
        bot.register_next_step_handler_by_chat_id(user,partial(confirm_this_fio))

    cur.close()
    con.close()


#@bot.callback_query_handler(func=lambda call: call.data == 'make_order')
#def handle_make_order(call):
#    bot.send_message(call.message.chat.id, f"Наберите, что бы вы хотели заказать {random.choice(typing_emojis)}")
#    bot.register_next_step_handler_by_chat_id(call.from_user.id, partial(process_order))

# Функция для обработки введённого заказа
def process_order(message):
    order_text = message.text
    user = message.chat.id

    # Создание инлайн-клавиатуры с кнопками "Да, отправить" и "Нет, изменить"
    markup = types.InlineKeyboardMarkup()
    yes_button = types.InlineKeyboardButton(f"Да, отправить {random.choice(send_order_emojis)}", callback_data='confirm_order')
    no_button = types.InlineKeyboardButton(f"Нет, изменить {random.choice(delay_order_emojis)}", callback_data='change_order')
    markup.add(yes_button, no_button)
    user_order[user] = str()
    user_order[user] = order_text
    # Сохранение заказа во временное хранилище (например, в user_data)
    bot.send_message(user, f"Отправить ваш заказ? {random.choice(order_emojis)}", reply_markup=markup)
    #bot.register_next_step_handler(message, lambda msg: confirm_order(msg, order_text))

# Обработчик подтверждения заказа
@bot.callback_query_handler(func=lambda call:True and call.data =='confirm_order')
def confirm_order(call:CallbackQuery):
    user = call.from_user.id
    con = sqlite3.connect(db)
    cur = con.cursor()
    order_text = user_order[user]
    # Сохранение заказа в базе данных
    cur.execute(f'''UPDATE users SET order_text='{order_text}' WHERE id ={user} ''')
    
    con.commit()
    arc = cur.execute(f'''SELECT archive FROM users WHERE id = {user} ''').fetchone()
    if arc is not None:
        if arc[0] is not None:
            new_arc = arc+'-'+order_text
            
            cur.execute(f'''UPDATE users SET archive='{new_arc}' WHERE id ={user} ''')
    
            con.commit()
        else:
            cur.execute(f'''UPDATE users SET archive='{order_text}' WHERE id ={user} ''')
    
            con.commit()
    else:
        cur.execute(f'''UPDATE users SET archive='{order_text}' WHERE id ={user} ''')
    
        con.commit()


    if call.data =='confirm_order':
        bot.send_message(user, f"Ваш заказ отправлен!{random.choice(order_emojis)}\nВ течение часа ожидайте ответа)")
        new_order = order_text.replace('-','')
        # Уведомление администратора о новом заказе
        bot.send_message(admin_id, f"Новый заказ\n\'{new_order}\'")
        phone = cur.execute(f'''SELECT phone FROM users WHERE id = {user} ''').fetchone()[0]
        n = cur.execute(f'''SELECT ФИО FROM users WHERE id = {user} ''').fetchone()[0]
        full_name = n.split()
        full_name = full_name[0]+" "+full_name[1]
        do = telebot.types.InlineKeyboardMarkup()
        d = telebot.types.InlineKeyboardButton(f'Выполнить заказ {random.choice(order_emojis)}',callback_data=f'do_it_{user}')
        bot.send_message(admin_id,f'От: {full_name}\nТелефон{random.choice(phone_emojis)}:{phone}')

        # Подтверждение пользователю
        

    cur.close()
    con.close()
    
# Обработчик изменения заказа
@bot.callback_query_handler(func=lambda call: call.data == 'change_order')
def handle_change_order(call:CallbackQuery):
    user = call.from_user.id
    con = sqlite3.connect(db)
    cur = con.cursor()
    cur.execute(f'''UPDATE users SET order_text=NULL WHERE id ={user} ''')
    
    con.commit()
    bot.send_message(call.message.chat.id, "Пожалуйста, введите ваш новый заказ:")
    user_order[user] = ''
    bot.register_next_step_handler_by_chat_id(call.from_user.id,partial(process_order))
    

    cur.close()
    con.close()
    

@bot.callback_query_handler(func=lambda call: call.data.startswith('do_it'))
def doitorder(call:CallbackQuery):
    con = sqlite3.connect(db)
    cur = con.cursor()
    that_user = int(call.data.split('_')[-1])
    that_order = cur.execute(f'''SELECT order_text FROM users WHERE id ={that_user} ''').fetchone()
    n = cur.execute(f'''SELECT ФИО FROM users WHERE id ={that_user} ''').fetchone()[0]
    p = cur.execute(f'''SELECT phone FROM users WHERE id ={that_user} ''').fetchone()[0]
    if that_order is not None:
        if that_order[0] is not None:
            prep = cur.execute(f'''SELECT prepay FROM users WHERE id = {that_user} ''').fetchone()[0]
            if prep == 'False':
                bot.send_message(admin_id,'Учтите, предоплату пользователь еще не собирался вводить либо не уведомил вас об этом')
                ss = f'<a href="tg://user?id={that_user}">{n}</a>'
                bot.send_message(admin_id,f'Проверьте поступления на ваш банк от {n} ({p})\nСсылка: {ss} ',reply_markup=prove_prepay(that_user))
    
            bot.send_message(admin_id,'Отлично!\nВы взялись за заказ)',reply_markup=ReplyKeyboardRemove())
            bot.send_message(admin_id,'Главное меню',reply_markup=admin_markup())
            bot.clear_step_handler_by_chat_id(admin_id)
            cur.execute(f'''UPDATE users SET order_text = NULL WHERE id ={that_user} ''')
            con.commit()
        else:
            bot.send_message(admin_id,'Заказ уже в вашей обработке)')
            #order = cur.execute(f'''SELECT archive FROM users WHERE id ={that_user} ''').fetchone()
            #phone = cur.execute(f'''SELECT phone FROM users WHERE id ={that_user} ''').fetchone()
            #name = cur.execute(f'''SELECT ФИО FROM users WHERE id ={that_user} ''').fetchone()
            #if order is not None:
            #    if order[0] is not None:
            #        o = order[0].split('-')[-1]
            #        msg = f'От: {name[0]}\nЗаказ{random.choice(order_emojis)}: {o}\nТелефон{random.choice(phone_emojis)}: {phone[0]}'
            #        bot.send_message(admin_id,f"Напоминаем:\n{msg}")
            bot.send_message(admin_id,'Главное меню',reply_markup=admin_markup())




    cur.close()
    con.close()


@bot.callback_query_handler(func=lambda call: call.data == 'send_nomen')
def nomenclatura(call:CallbackQuery):
    con = sqlite3.connect(db)
    cur = con.cursor()

    bot.send_message(admin_id,'Отправьте Экселевский файл, я его считаю)',reply_markup=glav())
    bot.clear_step_handler_by_chat_id(admin_id)
    bot.register_next_step_handler_by_chat_id(admin_id,read_excel)


def read_excel(message:Message):
    con = sqlite3.connect(db)
    cur = con.cursor()
    global categories
    if not message.document:
        bot.send_message(admin_id,'Отправьте экселевский файл)',reply_markup=glav())
        bot.clear_step_handler_by_chat_id(admin_id)
        bot.register_next_step_handler_by_chat_id(admin_id,read_excel)
    else:
        try:
        # Получение информации о документе
            
            file_info = bot.get_file(message.document.file_id)
            file_name = message.document.file_name.lower()

            # Проверка расширения файла
            if file_name.endswith(('.xls', '.xlsx')):
                # Скачивание файла
                downloaded_file = bot.download_file(file_info.file_path)
                #file_path = os.path.join('downloads', message.document.file_name)

                # Сохранение файла на сервере
                file_path = '/var/www/Bakery/Catalog.xlsx'
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active

                categories = {}
                d = 0
        
            
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=4, values_only=True):
            
                    category, product, detail1, detail2 = row
            
                    
                    print("row ",row)
                    detail1 = str(row[2])
                    detail2 = str(row[3])
                    if category not in categories and category is not None and category!='':
                        print("Categorey: ",category)
                
                        categories[category] = {}
                        details = detail1 + ', '+ detail2
                        categories[category][product] =str(d+1)+', '+ details
                
                    elif category == '':
                        continue
                    elif category is None:
                        continue
                    elif category in categories:
                        if product not in categories[category].keys():
                    
                            print("Product: ",product," appended")
                            categories[category][product] = str(d+1)+', '+ detail1 + ','+ detail2
                    
                            
                    d = d+1
                workbook.close()
                c = 0
                for k in categories.keys():
                    print("K: ",k)
                    c+=1
                    dict_of_cats[c]={}
                    dict_of_cats[c][k]={}
                    print("C: ",c)
            
                    dict_of_cats[c][k] = categories[k]
       
                bot.send_message(admin_id,'Каталог загружен)',reply_markup=admin_markup())
                bot.clear_step_handler_by_chat_id(admin_id)
            else:
                bot.send_message(admin_id,'Отправьте экселевский файл)',reply_markup=glav())
                bot.clear_step_handler_by_chat_id(admin_id)
                bot.register_next_step_handler_by_chat_id(admin_id,read_excel)
        except Exception as e:
            bot.send_message(admin_id,f'Возникла ошибка: {e}',reply_markup=ReplyKeyboardRemove())
            
            bot.send_message(admin_id,'Обратитесь к разработчику')
            bot.send_message(admin_id,'Главное меню',reply_markup=admin_markup())

    cur.close()
    con.close()



@bot.callback_query_handler(func=lambda call :True and call.data.startswith('make_order'))
def show_categories(call:CallbackQuery):
    try:
        user = call.from_user.id
        global all_orders
        global dict_of_cats
        global categories
        global normal_orders
        print("CATEGORIES: ",categories)
        if user not in [keys for keys in all_orders.keys()]:
            all_orders[user] = []
    
        if user in [keys for keys in normal_orders.keys()]:
        
            if len(normal_orders[user]) != 0:
                for i in normal_orders[user]:
                    if i.startswith('adress_'):
                        normal_orders[user].remove(i)
                
        keyboard =telebot.types.InlineKeyboardMarkup()
    
        
    
    
        s = list()
        for i in dict_of_cats.keys():
            s.append(i)
        print("List of cats: ",s)
        row = []
    
    
        row = []
        for i, category in enumerate(categories):
            print("I: ",i)
            button_text = f"{category} {random.choice(bakery_emojis)}"
            callback_data = f'category_{(i+1)}'
            button = telebot.types.InlineKeyboardButton(text=button_text, callback_data=callback_data)
            row.append(button)
            # Если в строке три кнопки, добавляем её в клавиатуру и начинаем новую строку
            if len(row) == 3:
                keyboard.row(*row)
                row = []

    # Добавляем оставшиеся кнопки, если они есть
        if row:
            keyboard.row(*row)
    
    
        
        keyboard.add(telebot.types.InlineKeyboardButton(text='Выйти', callback_data='glav_from_order'))
        if user in [keys for keys in all_orders.keys()]:
            if len(all_orders[user])!=0:
                keyboard.row(telebot.types.InlineKeyboardButton(text=f'Отправить мой заказ {random.choice(order_emojis)}',callback_data='send_my_order'))
    
        bot.send_message(user,'Выберите категорию)',reply_markup=keyboard)
    except Exception as e:
        bot.send_message(1159187641,f'{e}',reply_markup=client_markup())



@bot.callback_query_handler(func=lambda call: call.data.startswith('category_'))
def show_products(call):
    integ = int(call.data.split('_')[-1])
    print("INTeg: ",integ)
    global categories
    print("Categories:",categories)
    global all_orders
    global dict_of_cats
    global normal_orders
    print("MAIN dict:",dict_of_cats)
    print("Dict of cats keys\n\n\n",dict_of_cats.keys())
    category = dict_of_cats[integ]
    for i in category.keys():
        print("Category,keys ",i)
        category = i
        print("Category = ",i)
    if category in categories:
        print("in")
        row = []
        keyboard = types.InlineKeyboardMarkup()
        for product in categories[category].keys():
            numba = dict_of_cats[integ][category][product].split(',')[-3]
            print(numba)
            print(product, 'in')
            button = types.InlineKeyboardButton(text=product, callback_data=f'product_{integ}_{numba}')
            row.append(button)
            if len(row) == 2:
                keyboard.row(*row)
                row = []
        if row:
            keyboard.row(*row)
        keyboard.add(telebot.types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data='make_order'))
        keyboard.add(telebot.types.InlineKeyboardButton(text='Выйти', callback_data='glav_from_order'))
        bot.send_message(chat_id=call.from_user.id,text=f'Продукты в категории {category}:', reply_markup=keyboard)
    else:
        print("Not in")
        



@bot.callback_query_handler(func=lambda call: call.data.startswith('product_'))
def show_details(call):
    global categories
    global all_orders
    global dict_of_cats
    global normal_orders
    user = call.from_user.id
    _, category, product = call.data.split('_', 2)
    numba = int(call.data.split('_')[-1])
    category = call.data.split('_')[-2]
    print("Numba: ",numba)
    
    try:
        index = categories[category].find(product)
        print("Index: ",index)
    except:
        index = 0
      
    integ = int(call.data.split('_')[-2])
    category = dict_of_cats[integ]
    flag = False
    for i in dict_of_cats[integ].keys():
        if flag == False:
            for j in dict_of_cats[integ][i].keys():
                values = dict_of_cats[integ][i][j].split(',')
                print("Values: ",values)
                if numba == int(values[-3]):
                    print("Found")
                    product = j
                    category = i
                    flag = True
                    break
            
                                                     
    
    '''for k in categories[category].keys():
        if int(categories[category][k].split(',')[0])==numba:
            print(categories[category][k].split(',')[0])
            product = k
            print("My product: ",product)
            break
        else:
            continue
    else:
        print("Nope")
        return'''
    product = str(product)
    photos = ['Торт Зебра.jpg', "Брауни-вишня-чизкейк.jpg",'Миндаль-грецкий орех.jpg','Трубочка с карамельным кремом.jpg']
    if "зебра" in product.lower():
        photo = photos[0]
    elif "вишня-чизкейк" in product.lower():
        photo = photos[1]
    elif 'миндаль-грецкий' in product.lower():
        photo = photos[2]
    elif 'с карамельным кремом' in product.lower():
        photo = photos[3]
    else:
        photo = 'none'

    if take_photo(product) == True:
        photo = product+'.jpg'
        
    try:
        c = all_orders[user].count(product)
    except:
        c = 0
    print("Product: ",product)
    print("Category: ",category)
    ismer = categories[category][product].split(',')[-2]
    print("Ismerenie: ",ismer)
    if photo == 'none':
        if c == 0:
            print("C = 0")
            if category in categories.keys() and product in categories[category].keys():
                if "порц" in ismer:
                    details = categories[category][product]
                    ism = details.split(',')[-2]
                    price = details.split(',')[-1]
                    detail_text = f"Продукт: {product}\nЕд. изм.: {ism}\nЦена: {price}руб./{ism}"
                    keyboard = telebot.types.InlineKeyboardMarkup()
            
                    button_back = types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data='make_order')
                    button_order = types.InlineKeyboardButton(text='Кол-во', callback_data=f'iord_{integ}_{numba}_por')
                    button_main_menu = types.InlineKeyboardButton(text='Выйт', callback_data='glav_from_order')
                    keyboard.row(button_back, button_order, button_main_menu)
                    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                            text=detail_text, reply_markup=keyboard)
            
                else:
                    details = categories[category][product]
                    ism = details.split(',')[-2]
                    price = details.split(',')[-1]
                    detail_text = f"Продукт: {product}\nЕд. изм.: {ism}\nЦена: {price}руб./{ism}"
                    keyboard = telebot.types.InlineKeyboardMarkup()
            
                    button_back = types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data='make_order')
                    button_order1 = types.InlineKeyboardButton(text='Кусок', callback_data=f'iord_{integ}_{numba}_1/4')
                    #button_order2 = types.InlineKeyboardButton(text='Заказать (500г)', callback_data=f'iorder_{product}_1/2')
                    button_order3 = types.InlineKeyboardButton(text='Целый', callback_data=f'iord_{integ}_{numba}_1')
                    button_main_menu = types.InlineKeyboardButton(text='Выйти', callback_data='glav_from_order')
                    keyboard.row(button_order1,button_order3)
                    keyboard.row(button_back,button_main_menu)
                    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                            text=detail_text, reply_markup=keyboard)
            
                

        else:
            print("C!= 0")
            if category in categories.keys() and product in categories[category].keys():
                if "порц" in ismer:
                    details = categories[category][product]
                    ism = details.split(',')[-2]
                    price = details.split(',')[-1]
                    detail_text = f"Продукт: {product}\nЕд. изм.: {ism}\nЦена: {price}руб./{ism}"
                    keyboard = telebot.types.InlineKeyboardMarkup()
                    keyboard.row_width=3
                    m = telebot.types.InlineKeyboardButton(text='-',callback_data=f'minus_{product}')
                    p = telebot.types.InlineKeyboardButton(text='+',callback_data=f'plus_{product}')
                    q = telebot.types.InlineKeyboardButton(text=f'{c}',callback_data=f'count_{product}')
                    back = telebot.types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data=f'make_order')
                    to_ord = telebot.types.InlineKeyboardButton(text='Заказать', callback_data=f'iorder_{product}_por')
                    g = telebot.types.InlineKeyboardButton(text='Выйти', callback_data='glav_from_order')
                    keyboard.row(m,p,q)
                    keyboard.row(back,to_ord,g)
                    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                          text=detail_text, reply_markup=keyboard)
                else:
                    details = categories[category][product]
                    ism = details.split(',')[-2]
                    price = details.split(',')[-1]
                    detail_text = f"Продукт: {product}\nЕд. изм.: {ism}\nЦена: {price}руб./{ism}"
                    keyboard = telebot.types.InlineKeyboardMarkup()
                    keyboard.row_width=3
                    m = telebot.types.InlineKeyboardButton(text='-',callback_data=f'minus_{product}')
                    p = telebot.types.InlineKeyboardButton(text='+',callback_data=f'plus_{product}')
                    q = telebot.types.InlineKeyboardButton(text=f'{c}',callback_data=f'count_{product}')
                    back = telebot.types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data=f'make_order')
                    to_ord = telebot.types.InlineKeyboardButton(text='Заказать', callback_data=f'iorder_{product}')
                    g = telebot.types.InlineKeyboardButton(text='Выйти', callback_data='glav_from_order')
                    keyboard.row(m,p,q)
                    keyboard.row(back,to_ord,g)
                    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                          text=detail_text, reply_markup=keyboard)
    elif photo != 'none':
        if c == 0:
            print("C = 0")
            if category in categories.keys() and product in categories[category].keys():
                if "порц" in ismer:
                    details = categories[category][product]
                    ism = details.split(',')[-2]
                    price = details.split(',')[-1]
                    
                    detail_text = f"Продукт: {product}\nЕд. изм.: {ism}\nЦена: {price}руб./{ism}"
                    #with open(photo, 'rb') as image_file:
                    #    bot.send_photo(chat_id=call.message.chat.id, photo=image_file,caption=detail_text)
                    keyboard = telebot.types.InlineKeyboardMarkup()
            
                    button_back = types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data='make_order')
                    button_order = types.InlineKeyboardButton(text='Кол-во', callback_data=f'iord_{integ}_{numba}_por')
                    button_main_menu = types.InlineKeyboardButton(text='Выйт', callback_data='glav_from_order')
                    keyboard.row(button_back, button_order, button_main_menu)
                    #bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                    #                        text=detail_text, reply_markup=keyboard)
                    with open(photo, 'rb') as image_file:
                        bot.send_photo(chat_id=call.message.chat.id, photo=image_file,caption=detail_text,reply_markup=keyboard)
            
                else:
                    details = categories[category][product]
                    ism = details.split(',')[-2]
                    price = details.split(',')[-1]
                    detail_text = f"Продукт: {product}\nЕд. изм.: {ism}\nЦена: {price}руб./{ism}"
                    #with open(photo, 'rb') as image_file:
                    #    bot.send_photo(chat_id=call.message.chat.id, photo=image_file,caption=detail_text)
                    keyboard = telebot.types.InlineKeyboardMarkup()
            
                    button_back = types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data='make_order')
                    button_order1 = types.InlineKeyboardButton(text='Кусок', callback_data=f'iord_{integ}_{numba}_1/4')
                    #button_order2 = types.InlineKeyboardButton(text='Заказать (500г)', callback_data=f'iorder_{product}_1/2')
                    button_order3 = types.InlineKeyboardButton(text='Целый', callback_data=f'iord_{integ}_{numba}_1')
                    button_main_menu = types.InlineKeyboardButton(text='Выйти', callback_data='glav_from_order')
                    keyboard.row(button_order1,button_order3)
                    keyboard.row(button_back,button_main_menu)
                    #bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                    #                        text=detail_text, reply_markup=keyboard)
                    with open(photo, 'rb') as image_file:
                        bot.send_photo(chat_id=call.message.chat.id, photo=image_file,caption=detail_text,reply_markup=keyboard)
        

@bot.callback_query_handler(func=lambda call: call.data == 'back_to_categories')
def back_to_categories(call:CallbackQuery):
    show_categories(call.from_user.id)

@bot.message_handler(func=lambda message:True and (message.text!='/start' and message.text!='start') and message.from_user.id != admin_id)
def handle(message:Message):
    user = message.from_user.id
    bot.send_message(user,'Чем могу помочь?)')
    bot.send_message(user,'Воспользуйтесь клавиатурой для контакта с пекарней)',reply_markup=client_markup())
    bot.clear_step_handler_by_chat_id(user)
    


@bot.callback_query_handler(func=lambda call: call.data.startswith('iord'))
def orderins_smth(call:CallbackQuery):
    user = call.from_user.id
    con = sqlite3.connect(db)
    cur = con.cursor()
    global categories
    global all_orders
    global order_price
    global normal_orders
    
    
    integ = int(call.data.split('_')[-3])
    pr = dict_of_cats[integ]
    norm = str(call.data.split('_')[-2])
    num = int(call.data.split('_')[-2])
    print("Product: ",pr)
    q = None
    flag = False
    for k in dict_of_cats[integ].keys():
       if flag == False:
           for key in dict_of_cats[integ][k]:
               values = (dict_of_cats[integ][k][key]).split(',')
               numba = int(values[0])
               if numba == num:
                   product_full = key
                   flag = True
                   break
    data = call.data.split('_')[-1]
    
    if data == '1':
        q = "кг"
    if data == '1/2':
        q = "500г"
    if data == '1/4':
        q = 'Кусок'
    overall = int()
    if data =='por':
        q = 1
    
    try:
        overall = all_orders[user].count(pr.replace(' ',''))
    except:
        overall = 0

    if user not in order_price:
        order_price[user] = 0
    if len(all_orders[user]) == 0:
        order_price[user] = 0
        
    
    category = None
    flag = False
    ism = None
    for key in categories.keys():
        if flag == False:
            for k in categories[key].keys():
                if k == product_full:
                    category = key
                    ism = categories[key][k].split(',')[-2]
                    print("Category = ",key)
                    print("product = ",product_full)
                    flag = True
                    break
            
    weight = call.data.split('_')[-1]
    
    if weight == '1':
        weight = 1.0
    elif weight == '1/2':
        weight = 0.5
    elif weight == '1/4':
        weight = 0.25
    elif weight == 'por':
        weight =1.0
    to_add = None
    if "порц" in ism:
        to_add = product_full.replace(' ','')
        print("To add: ",to_add)
        
    else:
        to_add = product_full.replace(' ','')+q
        print("To add: ",to_add)
    all_orders[user].append(product_full)
    b = ['Кусок','кг']
    #for i in b:
    #    if i in norm:
    #        norm = norm.replace(i,'')
    
    photos = ['Торт Зебра.jpg', "Брауни-вишня-чизкейк.jpg",'Миндаль-грецкий орех.jpg','Трубочка с карамельным кремом.jpg']
    if "зебра" in product_full.lower():
        photo = photos[0]
    elif "вишня-чизкейк" in product_full.lower():
        photo = photos[1]
    elif 'миндаль-грецкий' in product_full.lower():
        photo = photos[2]
    elif 'с карамельным кремом' in product_full.lower():
        photo = photos[3]
    else:
        photo = 'none'
        
    if take_photo(product_full) == True:
        print(product_full)
        photo = product_full+'.jpg'
        print(photo)
    else:
        print("false")
    try:
        overall = all_orders[user].count(product_full)
    except:
        overall = 0
    print("Weight: ",weight)
    print("Overall: ",overall)
    category = None
    flag = False
    for k in categories.keys():
        
        if flag == False:
           
            print("FLag false")
            for i in categories[k].keys():
                
                if i.strip().lower() == product_full.strip().lower():
                    print(f"Сравниваем '{i.strip().lower()}' с '{product_full.strip().lower()}'")
                    print("True: ",pr,'   ',i)
                    category = k
                    flag = True
                    print("TRUE FLAGG")
                    break
                else:
                    continue
            if flag == True:
                break
        #else:
        #    print("Flag trued")
        #    break
    
        
    
    else:
        print("Mistake")
        return
    if 'порц' in ism:
        price = float(categories[category][product_full].split(',')[-1])*weight
    elif 'порц' not in ism and q == 'Кусок':
        price = 0
    elif "порц" not in ism and q == 'кг':
        price = float(categories[category][product_full].split(',')[-1])
    print("Price for ",product_full," is ",price)
    if user not in [keys for keys in normal_orders.keys()]:
        normal_orders[user]=[]
    normal_orders[user].append(product_full)
        
    keyboard = types.InlineKeyboardMarkup()
    order_price[user] = order_price[user] + int(price)
# Создаём кнопки
    pr = product_full.replace(' ','')    
    button_minus = types.InlineKeyboardButton(text='-', callback_data=f'minus_{integ}_{num}_{q}')
    button_count = types.InlineKeyboardButton(text=f'{overall}', callback_data=f'count')
    button_plus = types.InlineKeyboardButton(text='+', callback_data=f'plus_{integ}_{num}_{q}')
    button_send_order = types.InlineKeyboardButton(text=f'Отправьте мой заказ) {random.choice(order_emojis)}', callback_data='send_my_order')
    button_back = types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data='make_order')
    button_main_menu = types.InlineKeyboardButton(text='Выйти', callback_data='glav_from_order')

    # Добавляем кнопки в строки
    keyboard.row(button_minus, button_count, button_plus)
    
    keyboard.row(button_back)
    keyboard.row(button_send_order)
    keyboard.row(button_main_menu)
    if photo == 'none':
        bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,text=product_full,
                                   reply_markup=keyboard)

    else:
        with open(photo, 'rb') as new_image_file:
    # Создание объекта InputMediaPhoto
            
    
            # Редактирование сообщения
            bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id,reply_markup=keyboard)


    cur.close()
    con.close()
    


@bot.callback_query_handler(func=lambda call: call.data.startswith('minus') or call.data.startswith('plus'))
def orderins_smth(call:CallbackQuery):
    user = call.from_user.id
    global categories
    global all_orders
    global order_price
    global normal_orders
    c = int()
    pr = None
    integ = int(call.data.split('_')[-3])
    num = int(call.data.split('_')[-2])
    flag = False
    for i in dict_of_cats[integ].keys():
        if flag == False:
            for j in dict_of_cats[integ][i].keys():
                values = dict_of_cats[integ][i][j].split(',')
                numba = int(values[0])
                if num == numba:
                    pr = j
                    category = i
                    flag = True
                    break
                
    print("Call data: ",call.data)
    q = call.data.split('_')[-1].replace(' ','')
    if q == 'кг' or q =='1' or q == 1:
        q=1.0
    elif q == '500г':
        q = 2.0
    elif q == '250г':
        q = 4.0
    print("q:",q)    
    flag = False
    ism = None
    t = pr
    b = ['500г','250г','кг']
    #for i in b:
    #    if i in t:
    #        t = t.replace(i,'')
    #        print("T: ",t)
       
    if q in [1.0,'1.0']:
        for key in categories.keys():
            if flag == False:
                for k in categories[key].keys():
                    if k.replace(' ','')==t.replace(' ',''):
                        print(float(categories[key][k].split(',')[-1]))
                        price = float(categories[key][k].split(',')[-1])/float(q)
                        print(f"price:{price}")
                        flag = True
                        ism = categories[key][k].split(',')[-2]
                        break
                    else:
                        continue
    else:
        price = 0
        
        
    try:
        c = all_orders[user].count(pr)
        print("C:::: ",c)
    except Exception as e:
        bot.send_message(admin_id, e)
        print(e)
        all_orders[user] = []
        bot.send_message(user,'Возникла ошибка, администратор уже решает)',reply_markup=client_markup())
        return
    print("All orders: ",all_orders[user])
    
    photos = ['Торт Зебра.jpg', "Брауни-вишня-чизкейк.jpg",'Миндаль-грецкий орех.jpg','Трубочка с карамельным кремом.jpg']
    if "зебра" in pr.lower():
        photo = photos[0]
    elif "вишня-чизкейк" in pr.lower():
        photo = photos[1]
    elif 'миндаль-грецкий' in pr.lower():
        photo = photos[2]
    elif 'с карамельным кремом' in pr.lower():
        photo = photos[3]
    else:
        photo = 'none'
    d = None
    if len(call.data.split('_'))>2:
        d =call.data.split('_')[-1]
        
    norm = pr
    if call.data.startswith('minus'):
        c = c-1
        
        #pr = pr.replace(' ','')
        normal_orders[user].remove(norm)
        print("Pr: ",pr)
        
        
        all_orders[user].remove(pr)
        keyboard = telebot.types.InlineKeyboardMarkup()
        order_price[user] = float(order_price[user]) - float(price)
        keyboard = types.InlineKeyboardMarkup()
       
# Создаём кнопки
        
        
        print("Ism: ",ism)
        
        if q == 1:        
            if c != 0:        
                button_minus = types.InlineKeyboardButton(text='-', callback_data=f"minus_{integ}_{num}_{call.data.split('_')[-1].replace(' ','')}")
            button_count = types.InlineKeyboardButton(text=f'{c}', callback_data=f'count_')
            button_plus = types.InlineKeyboardButton(text='+', callback_data=f"plus_{integ}_{num}_{call.data.split('_')[-1].replace(' ','')}")
            button_send_order = types.InlineKeyboardButton(text=f'Отправьте мой заказ) {random.choice(order_emojis)}', callback_data='send_my_order')
            button_back = types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data='make_order')
            button_main_menu = types.InlineKeyboardButton(text='Выйти', callback_data='glav_from_order')

        # Добавляем кнопки в строки
            if c != 0:
                keyboard.row(button_minus, button_count, button_plus)
            else:
                keyboard.row(button_count, button_plus)
            keyboard.row(button_back)
            keyboard.row(button_send_order)
        
            keyboard.row(button_main_menu)
            if photo == 'none':
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,text=pr+f' ({c})',
                                        reply_markup=keyboard)
            else:
                bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                        reply_markup=keyboard)
                
        else:
            if c != 0:        
                button_minus = types.InlineKeyboardButton(text='-', callback_data=f"minus_{integ}_{num}_{call.data.split('_')[-1].replace(' ','')}")
            button_count = types.InlineKeyboardButton(text=f'{c}', callback_data=f"count_{call.data.split('_')[-1].replace(' ','')}")
            button_plus = types.InlineKeyboardButton(text='+', callback_data=f"plus_{integ}_{num}_{call.data.split('_')[-1].replace(' ','')}")
            button_send_order = types.InlineKeyboardButton(text=f'Отправьте мой заказ) {random.choice(order_emojis)}', callback_data='send_my_order')
            button_back = types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data='make_order')
            button_main_menu = types.InlineKeyboardButton(text='Выйти', callback_data='glav_from_order')

        # Добавляем кнопки в строки
            if c != 0:
                keyboard.row(button_minus, button_count, button_plus)
            else:
                keyboard.row(button_count, button_plus)
            keyboard.row(button_back)
            keyboard.row(button_send_order)
        
            keyboard.row(button_main_menu)
            if photo == 'none':
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,text=pr+f' ({c})',
                                        reply_markup=keyboard)
            else:
                bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                        reply_markup=keyboard)
            
        
    else:
        print("Pr: ",pr)
        normal_orders[user].append(norm)
        all_orders[user].append(pr)
        c = c+1
        keyboard = telebot.types.InlineKeyboardMarkup()
        order_price[user] = float(order_price[user]) + float(price)
        button_minus = types.InlineKeyboardButton(text='-', callback_data=f"minus_{integ}_{num}_{call.data.split('_')[-1].replace(' ','')}")
        button_count = types.InlineKeyboardButton(text=f'{c}', callback_data=f'count_')
        button_plus = types.InlineKeyboardButton(text='+', callback_data=f"plus_{integ}_{num}_{call.data.split('_')[-1].replace(' ','')}")
        back = telebot.types.InlineKeyboardButton(text=f'В меню {random.choice(nomenclature_emojis)}', callback_data=f'make_order')
        send_goddamn = telebot.types.InlineKeyboardButton(text=f'Отправьте мой заказ) {random.choice(order_emojis)}', callback_data=f'send_my_order')
        men = telebot.types.InlineKeyboardButton(text='Выйти', callback_data='glav_from_order')
        keyboard.row(button_minus, button_count, button_plus)
        keyboard.row(back)
        keyboard.row(send_goddamn)
        keyboard.row(men)
        if photo == 'none':
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,text=pr+f' ({c})',
                                        reply_markup=keyboard)
        else:
            bot.edit_message_reply_markup(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                    reply_markup=keyboard)

       
            



@bot.callback_query_handler(func=lambda call: call.data == 'send_my_order')
def sendingorder(call:CallbackQuery):
    user = call.from_user.id
    sam = 1
    ne_sam = 0
    dostavka = telebot.types.InlineKeyboardMarkup()
    s = telebot.types.InlineKeyboardButton(text='Заберу сам\n(самовывоз)',callback_data=f'dostavka_{sam}')
    n_s = telebot.types.InlineKeyboardButton(text='Оформить доставку (300 руб.)',callback_data=f'sam_{ne_sam}')
    v_menu = telebot.types.InlineKeyboardButton(text='В каталог',callback_data='make_order')
    otm = telebot.types.InlineKeyboardButton(text='Отменить заказ',callback_data='glav_from_order')
    dostavka.add(s,n_s)
    dostavka.add(v_menu)
    dostavka.add(otm)
    bot.send_message(user,'Как бы вы хотели забрать заказ?',reply_markup=dostavka)
    bot.clear_step_handler_by_chat_id(user)

@bot.callback_query_handler(func=lambda call: call.data.startswith('sam'))
def adress(call:CallbackQuery):
    global all_orders
    global order_price
    global normal_orders
    user = call.from_user.id
    bot.send_message(call.from_user.id,'Пожалуйста, введите адрес доставки:')
    bot.clear_step_handler_by_chat_id(user)
    bot.register_next_step_handler_by_chat_id(user,partial(enter_adress))

def enter_adress(message:Message):
    user = message.from_user.id
    yep_nope = telebot.types.InlineKeyboardMarkup()
    y = telebot.types.InlineKeyboardButton(text='Отправить адрес!',callback_data=f'dostavka_0')
    n = telebot.types.InlineKeyboardButton(text='Нет, исправить',callback_data=f'change_adress')
    v_menu = telebot.types.InlineKeyboardButton(text='В каталог',callback_data='make_order')
    otm = telebot.types.InlineKeyboardButton(text='Отменить заказ',callback_data='glav_from_order')
    yep_nope.add(y,n)
    yep_nope.add(v_menu)
    yep_nope.add(otm)
    adres = 'adress_'+message.text
    global all_orders
    global order_price
    global normal_orders
    if len(normal_orders[user]) != 0:
        normal_orders[user].append(adres)
    else:
        bot.send_message(user,'У вас еще нет заказов)',reply_markup=client_markup())
        bot.clear_step_handler_by_chat_id(user)
        return
    bot.send_message(message.from_user.id,f'Отправить данные?\n"{message.text}"',reply_markup=yep_nope)
    bot.clear_step_handler_by_chat_id(user)
    

@bot.callback_query_handler(func=lambda call: call.data == 'change_adress')
def adress(call:CallbackQuery):
    user = call.from_user.id
    global all_orders
    global order_price
    global normal_orders
    if len(normal_orders[user]) != 0:
        for i in normal_orders[user]:
            if i.startswith('adress_'):
                bot.send_message(user,'Пожалуйста, введите другой адрес')
                bot.clear_step_handler_by_chat_id(user)
                bot.register_next_step_handler_by_chat_id(user,partial(enter_adress))
                normal_orders[user].remove(i)
                break
        else:
            bot.send_message(user,'Кажется, вы ещё не вводили адрес\nВведите ваш адрес для доставки')
            bot.clear_step_handler_by_chat_id(user)
            bot.register_next_step_handler_by_chat_id(user,partial(enter_adress))
    else:
        bot.send_message(user,'У вас еще нет заказов)',reply_markup=client_markup())
        bot.clear_step_handler_by_chat_id(user)
        return


@bot.callback_query_handler(func=lambda call: call.data.startswith('dostavka'))
def finale(call:CallbackQuery):
    global all_orders
    global order_price
    global normal_orders
    def split_by_keywords(s, keywords):
        # Создаём регулярное выражение, объединяя ключевые слова
        pattern = '|'.join(map(re.escape, keywords))
        # Разделяем строку по шаблону
        parts = re.split(f'({pattern})', s)
        # Удаляем пустые строки из результата
        return [part for part in parts if part]
    user = call.from_user.id
    con = sqlite3.connect(db)
    cur = con.cursor()
    unq = set()
    unq = list(set(all_orders[user]))
    dict_o = {}
    print(all_orders[user])
    for i in unq:
        c = all_orders[user].count(i)
        dict_o[i] = c

    order_t = str()
    for i in dict_o.keys():
        order_text = cur.execute(f'''SELECT order_text FROM users WHERE id = {user} ''').fetchone()
        if order_text is not None:
            keywords = ['Кусок', 'кг']
            if order_text[0] is not None:
                
                for s in keywords:
                     if s in i:
                         print(s,' in ',i)
                         #result = split_by_keywords(i,keywords)
                         r = i.replace(s,'')
                         r = r+" "+s
                         print("r: ",r)
                         order_text = order_text[0] + f'{r}: {dict_o[i]}\n'
                         order_t =order_t+ order_text[0] + f'{r}: {dict_o[i]}\n'
                         break
                else:
                    print("no s in i")
                    order_text = order_text[0] + f'{i}: {dict_o[i]}\n'
                    order_t =order_t+ order_text[0] + f'{i}: {dict_o[i]}\n'
            else:
                for s in keywords:
                     if s in i:
                         print(s,' in ',i)
                         #result = split_by_keywords(i,keywords)
                         r = i.replace(s,'')
                         
                         r = r+" "+s
                         print("r: ",r)
                         order_text =f'{r}: {dict_o[i]}\n'
                         order_t =order_t+f'{r}: {dict_o[i]}\n'
                         break
                else:
                    print("no s in i")
                    order_text =f'{i}: {dict_o[i]}\n'
                    order_t =order_t+ f'{i}: {dict_o[i]}\n'
                
                
            
                

        cur.execute(f'''UPDATE users SET order_text = '{order_text}' WHERE id ={user}''')
        con.commit()
        print("order t = ",order_t)
    
    normal = ''
    nr = {}
    adres = ''
    for s in normal_orders[user]:
        if not s.startswith('adress_'):
            c = normal_orders[user].count(s)
            nr[s] = c
        else:
            adres = s.split('_')[-1]
    for k, v in nr.items():
        normal += k+" : "+str(v)+'\n'
        
    total = order_price[user]
    o = cur.execute(f'''SELECT order_text FROM users WHERE id = {user} ''').fetchone()[0]
    name = cur.execute(f'''SELECT ФИО FROM users WHERE id = {user} ''').fetchone()[0]
    phone = cur.execute(f'''SELECT phone FROM users WHERE id = {user} ''').fetchone()[0]
    on_the_way = int(call.data.split('_')[-1])
    if on_the_way in ['1',1]:
        on_the_way = 'Самовывоз'
        
    else:
        on_the_way = 'Доставка из пекарни'
        total = total+300
    d = telebot.types.InlineKeyboardMarkup()
    v = telebot.types.InlineKeyboardButton(text='Выполнить заказ!',callback_data=f'do_it_{user}')
    d.add(v)
    
    if total != 0:
        if order_price[user] not in [0,'0']:
            bot.send_message(user,f'Ваш заказ стоимостью {total} рублей успешно отправлен в пекарню)\nОжидайте ответа в течение часа')
        elif order_price[user] in [0,'0'] and total in [300,'300']:
            bot.send_message(user,f'В стоимость заказа вошла доставка ({total} руб.)\nОжидайте ответа в течение часа')
    else:
        bot.send_message(user,f'За кусок торта предоплата не полагается)\nОжидайте ответа в течение часа')
    if adres != '':
        bot.send_message(user,text=f'От {random.choice(booking_emojis)}: {name}\nЗаказ {random.choice(order_emojis)}: {normal}\nСпособ доставки:{on_the_way} на {adres}')
    else:
        bot.send_message(user,text=f'От {random.choice(booking_emojis)}: {name}\nЗаказ {random.choice(order_emojis)}: {normal}\nСпособ доставки:{on_the_way}')
    users_order = cur.execute(f'''SELECT orders_all FROM users WHERE id = {user} ''').fetchone()
    if users_order is not None:
        if users_order[0] is not None:
            users_orders = users_order[0]+'__'+normal
            print("ORDER: ",users_orders)
            cur.execute(f'''UPDATE users SET orders_all = '{users_orders}' WHERE id = {user} ''')
            con.commit()
        else:
            
            cur.execute(f'''UPDATE users SET orders_all = '{normal}' WHERE id = {user} ''')
            con.commit()
    #bot.send_message(user,text=f'Обязательно внесите предоплату 50% по номеру: {89139270431} на сбербанк',reply_markup=pred(user))
    fls = 'False'
    cur.execute(f'''UPDATE users SET prepay = '{fls}' WHERE id = {user} ''')
    con.commit()
    order_t = cur.execute(f'''SELECT order_text FROM users WHERE id = {user} ''').fetchone()
    '''if total != 0:
        thread = threading.Thread(target=predoplata, args=(user,total,order_t,))
        thread.start()
    else:
        bot.send_message(user,'Главное меню',reply_markup=client_markup())
        bot.clear_step_handler_by_chat_id(user)'''
    n = 'Оксана Геннадьевна'
    s = f'<a href="tg://user?id={admin_id}">{n}</a>'
    if total!=0:
        bot.send_message(user,f'Пожалуйста, внесите 50% ({total/2}) для вашего заказа по номеру: 89138270431\nБанк: Сбербанк\nПолучатель: {s}',parse_mode='HTML',reply_markup=client_markup()) 
      
    else:
        bot.send_message(user,'Заказ передан на оформление',reply_markup=client_markup())
    #bot.send_message(user,'Главное меню',reply_markup=client_markup())
    bot.send_message(admin_id,'Новый заказ!')
    if adres == '':
        bot.send_message(admin_id,f'От {random.choice(booking_emojis)}: {name}\nЗаказ {random.choice(order_emojis)}: {o}\nТелефон{random.choice(phone_emojis)}: {phone}\nПредоплата ожидается\nСпособ доставки:{on_the_way}.',reply_markup=d)
    else:
        bot.send_message(admin_id,f'От {random.choice(booking_emojis)}: {name}\nЗаказ {random.choice(order_emojis)}: {o}\nТелефон{random.choice(phone_emojis)}: {phone}\nПредоплата ожидается\nСпособ доставки:{on_the_way} на {adres}',reply_markup=d)

    bot.send_message(admin_id,'Главное меню',reply_markup=admin_markup())
    cur.execute(f'''UPDATE users SET order_text=NULL WHERE id = {user} ''')
    con.commit()
    order_price[user] = 0
    normal_orders[user] = []
    all_orders[user] = []
    #cur.execute(f'''UPDATE users SET order_text = NULL WHERE id ={user} ''')
    #con.commit()
    cur.close()
    con.close()




@bot.callback_query_handler(func=lambda call: True and call.data == 'glav_from_order')
def v_glavnoe(call:CallbackQuery):
    con = sqlite3.connect(db)
    cur = con.cursor()
    global all_orders
    user = call.from_user.id
    if len(all_orders[user])!=0:
        all_orders[user] = []
    cur.execute(f'''UPDATE users SET order_text = NULL WHERE id = {call.from_user.id} ''')
    con.commit()
    bot.send_message(call.from_user.id,'Возвращаемся)',reply_markup=ReplyKeyboardRemove())
    bot.send_message(call.from_user.id,'Главное меню',reply_markup=client_markup())
    bot.clear_step_handler_by_chat_id(call.from_user.id)

    cur.close()
    con.close()




@bot.callback_query_handler(func=lambda call: call.data == 'quest')
def checking(call:CallbackQuery):
    user = call.from_user.id
    con = sqlite3.connect(db)
    cur = con.cursor()
    phone = cur.execute(f'''SELECT phone FROM users WHERE id = {admin_id} ''').fetchone()[0]
    msg = f'{phone}\nОксана Геннадьевна'
    s = f'<a href="tg://user?id={admin_id}">{msg}</a>'
    message_text = (
    'Введите ваш вопрос или свяжитесь напрямую по телефону: '
    f'{s}'
)
    bot.send_message(user, message_text, parse_mode='HTML', reply_markup=glav())
    bot.clear_step_handler_by_chat_id(user)
    bot.register_next_step_handler_by_chat_id(user,partial(asking))

    cur.close()
    con.close()


def asking(message:Message):
    user = message.from_user.id
    con = sqlite3.connect(db)
    cur = con.cursor()
    cur.execute(f'''UPDATE users SET questions = '{message.text}' WHERE id = {user} ''')
    con.commit()
    yes_or_no = telebot.types.InlineKeyboardMarkup()
    y = telebot.types.InlineKeyboardButton(text='Да, отправить!',callback_data='send_mes')
    n = telebot.types.InlineKeyboardButton(text='Нет, исправить',callback_data="not_send")
    yes_or_no.add(y,n)
                                           
    bot.send_message(user,f'Отправляем?\n{message.text}',reply_markup=yes_or_no)

    cur.close()
    con.close()


@bot.callback_query_handler(func=lambda call: True and call.data == 'glav')
def v_glavnoe(call:CallbackQuery):
    user = call.from_user.id
    global msg
    if user != admin_id:
        bot.send_message(user,'Главное меню',reply_markup=client_markup())
        bot.clear_step_handler_by_chat_id(user)

    else:
        try:
            msg.__del__()
        except:
            pass
        bot.send_message(user,'Главное меню',reply_markup=admin_markup())
        bot.clear_step_handler_by_chat_id(user)





@bot.callback_query_handler(func=lambda call: call.data == 'send_mes' or call.data == 'not_send')
def mes_send_or_not(call:CallbackQuery):
    user = call.from_user.id
    con = sqlite3.connect(db)
    cur = con.cursor()
    if call.data == 'send_mes':
        bot.send_message(user,'Ваше сообщение отправлено!)',reply_markup=client_markup())
        mes = cur.execute(f'''SELECT questions FROM users WHERE id = {user} ''').fetchone()
        name = cur.execute(f'''SELECT ФИО FROM users WHERE id = {user} ''').fetchone()
        phone = cur.execute(f'''SELECT phone FROM users WHERE id = {user} ''').fetchone()
        if mes is not None:
            if mes[0] is not None:
                mes = mes[0]
                bot.send_message(admin_id,f'Новое обращение:\nОт:{name}\nТекст: {mes}\nТелефон:{phone}')
                cur.execute(f'''UPDATE users SET questions = NULL WHERE id = {user} ''')
                con.commit()
            else:
                bot.send_message(user,'Возникла какая-то ошибка\nАдминистратор уже решает',reply_markup=client_markup())
                bot.clear_step_handler_by_chat_id(user)
                return
        else:
            bot.send_message(user,'Возникла какая-то ошибка\nАдминистратор уже решает',reply_markup=client_markup())
            bot.clear_step_handler_by_chat_id(user)
            return
        
        bot.clear_step_handler_by_chat_id(user)
        
    elif call.data == 'not_send':
        bot.send_message(user,'Исправьте ваше особщение)',reply_markup=glav())
        cur.execute(f'''UPDATE users SET questions = NULL WHERE id = {user} ''')
        con.commit()
        bot.clear_step_handler_by_chat_id(user)
        bot.register_next_step_handler_by_chat_id(user,partial(asking))
                                           
    
    cur.close()
    con.close()



@bot.callback_query_handler(func=lambda call: call.data == 'check_orders' and call.from_user.id==admin_id)
def checking(call:CallbackQuery):
    con = sqlite3.connect(db)
    cur = con.cursor()
    now = datetime.datetime.now()
    data = []
    hat = ["ФИО", "Телефон","Заказ"]
    data.append(hat)
    con.text_factory = lambda b: b.decode('utf-8','ignore')
    all_users = cur.execute(f'''SELECT id FROM users ''').fetchall()
    if all_users:
        if all_users is not None:
            for i in all_users:
                if i[0] != admin_id:
            
                    this_user = []
                    fio = cur.execute(f'''SELECT ФИО FROM users WHERE id = {i[0]}''').fetchone()[0]
                    order = cur.execute(f'''SELECT orders_all FROM users WHERE id = {i[0]}''').fetchone()[0]
                    phone = cur.execute(f'''SELECT phone FROM users WHERE id = {i[0]}''').fetchone()[0]
                    fio = fio.encode("latin1").decode("utf-8") if isinstance(fio, bytes) else fio
                    phone = phone.encode("latin1").decode("utf-8") if isinstance(phone, bytes) else phone
                    order = order.encode("latin1").decode("utf-8") if isinstance(order, bytes) else order
                    try:
                        order = order.split('__')
                    except:
                        pass
                    print("order: ",order)
                    this_user.append(fio)
                    this_user.append(phone)
                    whole_order = str()
                    if order is not None:
                        if order !='None':
                            for f in order:
                                whole_order += f+'\n'
                        else:
                            whole_order = ''
                    else:
                        whole_order = ''
                    this_user.append(whole_order)
                    data.append(this_user)
                    
            print("data: ",data)
            pdf_filename = f"Заказы за сегодня.pdf"
            pdfmetrics.registerFont(ttfonts.TTFont('Arial', f'{arial}'))

            today_date = datetime.datetime.now().strftime("%Y-%m-%d")
            pdf_file = "Заказы за сегодня.pdf"
            
            c = SimpleDocTemplate(pdf_file, pagesize=letter)
            styles = getSampleStyleSheet()
            title = f"Все заказы за {today_date}"
            title_style = styles['Title']
            title_style.fontName = 'Arial'
            title_style.fontSize = 16
            table = Table(data)
            print("table")
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Arial'), 
                ('FONTNAME', (0, 1), (-1, -1), 'Arial'),  
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            print("elements")
            
            elements = [table]
            
            print("appended")
            c.build(elements)

            print(f"Файл {pdf_filename} создан успешно!")
            with open(pdf_file, "rb") as file:
                
                bot.send_document(admin_id,document=file,caption=F'Заказы за {today_date}')
            bot.send_message(admin_id,'Главное меню',reply_markup=admin_markup())
        else:
            bot.send_message(admin_id,'Кажется, на настоящий момент нет заказов)',reply_markup=admin_markup())
    else:
        bot.send_message(admin_id,'Кажется, на настоящий момент нет заказов)',reply_markup=admin_markup())
    
                


    cur.close()
    con.close()
    

@bot.callback_query_handler(func=lambda call: call.data == 'notif' and call.from_user.id==admin_id)
def notss(call:CallbackQuery):
    global msg
    msg = Notification()
    bot.send_message(admin_id,"Введите текст, аудио или видео (фото или документ)",reply_markup=glav())
    bot.register_next_step_handler_by_chat_id(admin_id,partial(nots))
    




def nots(message:Message):
    con = sqlite3.connect(db)
    cur = con.cursor()
    global msg
    users = cur.execute("SELECT id FROM users").fetchall()
    user_ids = [user[0] for user in users]
    #for user_id in user_ids:
    try:
        missing_elements = []

        if message.content_type == 'text':
            msg.text = message.text
        elif message.content_type in ['audio', 'voice']:
            msg.audio = message.audio.file_id if message.content_type == 'audio' else message.voice.file_id
        elif message.content_type == 'document':
            msg.document = message.document.file_id
        elif message.content_type == 'photo':
            msg.photo = message.photo[-1].file_id  # Берём последнюю (самую качественную) фотографию
        elif message.content_type == 'sticker':
            msg.sticker = message.sticker.file_id
        elif message.content_type == 'video':
            msg.video = message.video.file_id
        elif message.content_type == 'video_note':
            msg.video_note = message.video_note.file_id

        # Проверяем, какие элементы ещё не введены
        if not msg.text:
            missing_elements.append("текст")
        if not msg.audio:
            missing_elements.append("аудио")
        if not msg.document:
            missing_elements.append("документ")
        if not msg.photo:
            missing_elements.append("фото")
        if not msg.sticker:
            missing_elements.append("стикер")
        if not msg.video:
            missing_elements.append("видео")
        if not msg.video_note:
            missing_elements.append("видеозаметку")

        # Формируем сообщение для администратора
        if missing_elements:
            add_elements_text = "Добавить ещё элемент? Вот что ещё можно добавить: " + ", ".join(missing_elements)
        else:
            add_elements_text = "Все элементы добавлены)"

        # Отправляем сообщение администратору
        bot.send_message(admin_id, add_elements_text, reply_markup=add())
    except Exception as e:
        print(f"Не удалось отправить сообщение пользователю: {e}")

    cur.close()
    con.close()



@bot.callback_query_handler(func=lambda call:True and (call.data=='do_add' or call.data == 'not_add'))
def do_or_not(call:CallbackQuery):
    global msg
    to_send = [admin_id]
    if call.data == 'do_add':
        rechoice = telebot.types.InlineKeyboardMarkup()
        r = telebot.types.InlineKeyboardButton(text=f'Не добавлять {random.choice(skip_elements_emojis)}',callback_data='not_add')
        c = telebot.types.InlineKeyboardButton(text=f'Выйти {random.choice(main_menu_emojis)}',callback_data='otmena')
        otpr = telebot.types.InlineKeyboardButton(text='Так отправлю)',callback_data='do_send')
        rechoice.add(r,c,otpr)
        bot.send_message(admin_id,'Добавьте нужный элемент)',reply_markup=rechoice)
        bot.clear_step_handler_by_chat_id(admin_id)
        bot.register_next_step_handler_by_chat_id(admin_id,nots)
    if call.data == 'not_add':
        bot.send_message(admin_id,'Вот уведомление:')
        msg.send(bot,to_send)
        bot.send_message(admin_id,'Отправляем?',reply_markup=send_or_not())
        bot.clear_step_handler_by_chat_id(admin_id)



@bot.callback_query_handler(func=lambda call :True and call.data == 'otmena')
def otemintsend(call:CallbackQuery):
    global msg
    bot.send_message(admin_id,'В главное меню',reply_markup=admin_markup())
    bot.clear_step_handler_by_chat_id(admin_id)
    
    msg.__del__()
    


@bot.callback_query_handler(func=lambda call:True and (call.data=='do_send' or call.data == 'edit'))
def s_or_not(call:CallbackQuery):
    con = sqlite3.connect(db)
    cur = con.cursor()
    global msg
    users = cur.execute(f'''SELECT id FROM users WHERE id !={admin_id} ''').fetchall()
    
    if call.data == 'do_send':
        
        bot.send_message(admin_id,'Уведомление разослано)',reply_markup=admin_markup())
        bot.clear_step_handler_by_chat_id(admin_id)
        to_send = []
        for u in users:
            
            to_send.append(int(u[0]))
                #bot.send_message(u[0],'Уведомление от администратора!')
            
      
        try:
            
            msg.send(bot,to_send)
            for u in users:
                
                bot.send_message(int(u[0]),'Главное меню',reply_markup=client_markup())
        except Exception as e:
            print(e)
        
        #for u in users:
            
        #    bot.send_message(u[0],'Главное меню',reply_markup=client_markup())
    if call.data =='edit':      
        rechoice = telebot.types.InlineKeyboardMarkup()
        r = telebot.types.InlineKeyboardButton(text=f'Не добавлять {random.choice(skip_elements_emojis)}',callback_data='not_add')
        c = telebot.types.InlineKeyboardButton(text=f'В главное меню {random.choice(main_menu_emojis)}',callback_data='otmena')
        otpr = telebot.types.InlineKeyboardButton(text='Так отправлю)',callback_data='do_send')
        rechoice.add(r,c,otpr)
        bot.send_message(admin_id,"Введите нужный элемент (текст/видео/аудио и т.д.)",reply_markup=rechoice)
        bot.clear_step_handler_by_chat_id(admin_id)
        bot.register_next_step_handler_by_chat_id(admin_id,nots)
        
    cur.close()
    con.close()


def fetch_pending_orders():
    con = sqlite3.connect(db)
    cur = con.cursor()
    cur.execute("SELECT ФИО, phone, order_text FROM users WHERE id IS NOT ? AND order_text IS NOT NULL", (admin_id,))
    return cur.fetchall()

    cur.close()
    con.close()

# Function to notify admin about pending orders
def notify_admin():
    pending_orders = fetch_pending_orders()
    if pending_orders:
        # Create inline keyboard with "Посмотреть" button
        markup = types.InlineKeyboardMarkup()
        view_button = types.InlineKeyboardButton(f"Посмотреть {random.choice(view_emojis)}", callback_data='view_orders')
        markup.add(view_button)
        # Send notification to admin
        bot.send_message(admin_id, f"Есть еще невыполненные заказы {random.choice(order_emojis)}", reply_markup=markup)

# Function to run the notification loop
def notification_loop():
    while True:
        notify_admin()
        # Sleep for 3 hours (3 hours * 60 minutes * 60 seconds)
        time.sleep(1800)

# Start the notification loop in a separate thread
notification_thread = threading.Thread(target=notification_loop, daemon=True)
notification_thread.start()

# Handler for the "Посмотреть" button click
@bot.callback_query_handler(func=lambda call: call.data == 'view_orders')
def handle_view_orders(call):
    pending_orders = fetch_pending_orders()
    if pending_orders:
        for order in pending_orders:
            name, phone, order_text = order
            message = f"От: {name}\nЗаказ{random.choice(order_emojis)}: {order_text}\nТелефон{random.choice(phone_emojis)}: {phone}"
            bot.send_message(admin_id, message)
    else:
        bot.send_message(admin_id, "Нет невыполненных заказов)",reply_markup=admin_markup())
    bot.clear_step_handler_by_chat_id(admin_id)



def predoplata(user_id,total,order):
    # Функция для отправки напоминания о предоплате
    
    try:
        con= sqlite3.connect(db)
        cur = con.cursor()
        prepay = cur.execute(f'''SELECT prepay FROM users WHERE id ={user_id} ''').fetchone()
        
        if prepay is not None:
            if prepay[0] is not None:
                if prepay[0] == 'False':
                    prepay = False
                else:
                    prepay = True
                    return
        n = 'Оксана Геннадьевна'
        s = f'<a href="tg://user?id={admin_id}">{n}</a>'
        # Планирование напоминаний
        if total != 0:
            while prepay == False:
    
            # Отправка напоминания пользователю
                bot.send_message(user_id, f'Пожалуйста, внесите предоплату 50% ({total/2}) для вашего заказа по номеру: 89138270431\nБанк: Сбербанк\nПолучатель: {s}',parse_mode='HTML',reply_markup=client_markup_prepay(user_id))
            
                
                # Ожидание 2 часа перед следующим напоминанием
                time.sleep(60*60*2)
                prepay = cur.execute(f'''SELECT prepay FROM users WHERE id ={user_id} ''').fetchone()
                if prepay is not None:
                    if prepay[0] is not None:
                        if prepay[0] == 'False':
                            prepay = False
                        else:
                            bot.send_message(admin_id,f'{order}')
                            prepay = True
                            
        else:
            pass
        
    finally:
        cur.close()
        con.close()
    


@bot.callback_query_handler(func=lambda call: True and call.data.startswith('peredoplata'))
def pre(call:CallbackQuery):
    user = call.data.split('_')[-1]
    con = sqlite3.connect(db)
    cur = con.cursor()
    bot.send_message(user,f'Номер: {89138270431}\nБанк: Сбербанк\nПолучатель: Оксана Геннадьевна')
    bot.clear_step_handler_by_chat_id(user)
    bot.send_message(user,'Главное меню',reply_markup=client_markup())
    true = 'True'
    cur.execute(f'''UPDATE users SET prepay = '{true}' WHERE id = {user} ''')
    con.commit()
    name = cur.execute(f'''SELECT ФИО FROM users WHERE id = {user} ''').fetchone()[0]
    phone = cur.execute(f'''SELECT phone FROM users WHERE id = {user} ''').fetchone()[0]
      # Замените на реальный user_id
      # Замените на реальное имя пользователя
    full_msg = f'Имя: {name}\nТелефон:{phone}'
    # Создание сообщения с упоминанием пользователя
    message = f'<a href="tg://user?id={user}">{phone}</a>'

    # Отправка сообщения
    
    bot.send_message(admin_id,f'Пользователь {name} начал предоплату. Проверьте банк в ближайшее время)\nТелефон для связи: {message}',parse_mode='HTML')
    




    cur.close()
    con.close()
    





@bot.callback_query_handler(func=lambda call :True and call.data.startswith('pred_true'))
def true_pred(call:CallbackQuery):
    user = call.data.split('_')[-1]
    con = sqlite3.connect(db)
    cur = con.cursor()
    prepay = cur.execute(f'''SELECT prepay FROM users WHERE id = {user} ''').fetchone()
    if prepay is not None:
        if prepay[0] is not None:
            if prepay[0] == 'False':
                prepay = False
            else:
                prepay = True
        else:
            bot.send_message(admin_id,'Произошла какая-то ошибка\nОбратитесь к разработчику')
            bot.send_message(admin_id,'Главное меню',reply_markup=admin_markup())
            bot.clear_step_handler_by_chat_id(admin_id)
            return
        
    else:
        bot.send_message(admin_id,'Произошла какая-то ошибка в базе данных\nОбратитесь к разработчику')
        bot.send_message(admin_id,'Главное меню',reply_markup=admin_markup())
        bot.clear_step_handler_by_chat_id(admin_id)
        return

    if prepay == False:
        cur.execute(f'''UPDATE users SET prepay = 'True' WHERE id = {user} ''')
        con.commit()

        bot.send_message(user,'Администратор подтвердил вашу предоплату)')
    else:
        bot.send_message(admin_id,'Пользователь ввёл предоплату\nПроверьте банк!')
        bot.send_message(admin_id,'Главное меню',reply_markup=admin_markup())
        bot.clear_step_handler_by_chat_id(admin_id)


    cur.close()
    con.close()
    



def clear_orders():
    while True:
        now = datetime.datetime.now()
        #print("Сегодняшняя дата:", now.strftime("%Y-%m-%d"))
        month_names = {
    1: "января", 2: "февраля", 3: "марта", 4: "апреля",
    5: "мая", 6: "июня", 7: "июля", 8: "августа",
    9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
}
        formatted_date = f"{now.day} {month_names[now.month]}"
        # Проверяем, является ли текущее время 23:59
        if now.strftime("%H:%M:%S") == "23:59:00":
    
            con = sqlite3.connect(db)
            cur = con.cursor()
            data = []
            hat = ["ФИО", "Телефон","Заказ"]
            data.append(hat)
            all_users = cur.execute(f'''SELECT id FROM users ''').fetchall()
            if all_users:
                if all_users is not None:
                    for i in all_users:
                        if i[0] != admin_id:
            
                            this_user = []
                            fio = cur.execute(f'''SELECT ФИО FROM users WHERE id = {i[0]}''').fetchone()[0]
                            order = cur.execute(f'''SELECT orders_all FROM users WHERE id = {i[0]}''').fetchone()[0]
                            phone = cur.execute(f'''SELECT phone FROM users WHERE id = {i[0]}''').fetchone()[0]
                            fio = fio.encode("latin1").decode("utf-8") if isinstance(fio, bytes) else fio
                            phone = phone.encode("latin1").decode("utf-8") if isinstance(phone, bytes) else phone
                            order = order.encode("latin1").decode("utf-8") if isinstance(order, bytes) else order
                            try:
                                order = order.split('__')
                            except:
                                pass
                            print("order: ",order)
                            this_user.append(fio)
                            this_user.append(phone)
                            whole_order = str()
                            if order is not None:
                                if order !='None':
                                    for f in order:
                                        whole_order += f+'\n'
                                else:
                                    whole_order = ''
                            else:
                                whole_order = ''
                            this_user.append(whole_order)
                            data.append(this_user)
                    
                    print("data: ",data)
                    pdf_filename = f"Заказы за сегодня.pdf"
                    pdfmetrics.registerFont(ttfonts.TTFont('Arial', f'{arial}'))

                    today_date = datetime.datetime.now().strftime("%Y-%m-%d")
                    pdf_file = "Заказы за сегодня.pdf"
            
                    c = SimpleDocTemplate(pdf_file, pagesize=letter)
                    styles = getSampleStyleSheet()
                    title = f"Все заказы за {today_date}"
                    title_style = styles['Title']
                    title_style.fontName = 'Arial'
                    title_style.fontSize = 16
                    table = Table(data)
                    print("table")
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Arial'), 
                        ('FONTNAME', (0, 1), (-1, -1), 'Arial'),  
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    print("elements")
            
                    elements = [table]
            
                    print("appended")
                    c.build(elements)

                    print(f"Файл {pdf_filename} создан успешно!")
                    with open(pdf_file, "rb") as file:
                
                        bot.send_document(admin_id,document=file,caption=F'Заказы за {today_date}')
                    for i in all_users:
                        cur.execute(f'''UPDATE users SET orders_all = NULL WHERE id = {i[0]}''')
                        con.commit()
                    cur.close()
                    con.close()
        else:
            continue
        
#thr2 = threading.Thread(target=clear_orders, daemon=True)            
    

def clear_orders2():
    while True:
        now = datetime.datetime.now()
        #print("Сегодняшняя дата:", now.strftime("%Y-%m-%d"))
        month_names = {
    1: "января", 2: "февраля", 3: "марта", 4: "апреля",
    5: "мая", 6: "июня", 7: "июля", 8: "августа",
    9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
}
        formatted_date = f"{now.day} {month_names[now.month]}"
        # Проверяем, является ли текущее время 23:59
        if now.strftime("%H:%M:%S") == "23:59:00":
    
            con = sqlite3.connect(db)
            cur = con.cursor()
            data = []
            hat = ["ФИО", "Телефон","Заказ"]
            data.append(hat)
            all_users = cur.execute(f'''SELECT id FROM users ''').fetchall()
            if all_users:
                if all_users is not None:
                    for i in all_users:
                        if i[0] != admin_id:
            
                            this_user = []
                            fio = cur.execute(f'''SELECT ФИО FROM users WHERE id = {i[0]}''').fetchone()[0]
                            order = cur.execute(f'''SELECT orders_all FROM users WHERE id = {i[0]}''').fetchone()[0]
                            phone = cur.execute(f'''SELECT phone FROM users WHERE id = {i[0]}''').fetchone()[0]
                            fio = fio.encode("latin1").decode("utf-8") if isinstance(fio, bytes) else fio
                            phone = phone.encode("latin1").decode("utf-8") if isinstance(phone, bytes) else phone
                            order = order.encode("latin1").decode("utf-8") if isinstance(order, bytes) else order
                            try:
                                order = order.split('__')
                            except:
                                pass
                            print("order: ",order)
                            this_user.append(fio)
                            this_user.append(phone)
                            whole_order = str()
                            if order is not None:
                                if order !='None':
                                    for f in order:
                                        whole_order += f+'\n'
                                else:
                                    whole_order = ''
                            else:
                                whole_order = ''
                            this_user.append(whole_order)
                            data.append(this_user)
                    
                    print("data: ",data)
                    pdf_filename = f"Заказы за сегодня.pdf"
                    pdfmetrics.registerFont(ttfonts.TTFont('Arial', f'{arial}'))

                    today_date = datetime.datetime.now().strftime("%Y-%m-%d")
                    pdf_file = "Заказы за сегодня.pdf"
            
                    c = SimpleDocTemplate(pdf_file, pagesize=letter)
                    styles = getSampleStyleSheet()
                    title = f"Все заказы за {today_date}"
                    title_style = styles['Title']
                    title_style.fontName = 'Arial'
                    title_style.fontSize = 16
                    table = Table(data)
                    print("table")
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Arial'), 
                        ('FONTNAME', (0, 1), (-1, -1), 'Arial'),  
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    print("elements")
            
                    elements = [table]
            
                    print("appended")
                    c.build(elements)

                    print(f"Файл {pdf_filename} создан успешно!")
                    with open(pdf_file, "rb") as file:
                
                        bot.send_document(admin_id,document=file,caption=F'Заказы за {today_date}')
                    for i in all_users:
                        cur.execute(f'''UPDATE users SET orders_all = NULL WHERE id = {i[0]}''')
                        con.commit()
                    cur.close()
                    con.close()
        else:
            continue


if __name__ == "__main__":
    try:
        thr2 = threading.Thread(target=clear_orders2, daemon=True)
        thr2.start() 
        asyncio.run(main())
      

    except Exception as e:
        with open('/var/www/Bakery/logs.txt','w') as file:
            print("error", e)
    except KeyboardInterrupt:
        print('Exit')

bot.enable_save_next_step_handlers(delay=2)

# Load next_step_handlers from save file (default "./.handlers-saves/step.save")
# WARNING It will work only if enable_save_next_step_handlers was called!
bot.load_next_step_handlers()


